from win32com.client import Dispatch
import os
from PIL import ImageGrab
from openpyxl import load_workbook
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm

template_file = "Word成績報告模版.docx"
output_file   = "Word成績報告.docx"
output_pdf    = "Word成績報告.pdf"
output_img    = "成績圖表.png"
excel_file    = "成績管理4_barChart3D.xlsx"

app = Dispatch("Excel.Application")
app.Visible = 1
app.DisplayAlerts = 0
xlsx = app.Workbooks.Open(os.getcwd()+"/"+excel_file)
sheet = xlsx.Worksheets(1)
for x, chart in enumerate(sheet.Shapes):
    chart.Copy()
    image = ImageGrab.grabclipboard()
    image.save(output_img, "png")
    
xlsx.Close(False)
app.Quit()

wb = load_workbook(excel_file)
ws = wb.active
students = []
for row in range(2, ws.max_row + 1):
    data = {
             "name": ws.cell(row=row, column=1).value,
             "chinese": ws.cell(row=row, column=2).value,
             "english": ws.cell(row=row, column=3).value,
             "math": ws.cell(row=row, column=4).value
           }
    students.append(data)
wb.close()

doc = DocxTemplate(template_file)

img = InlineImage(doc, output_img, Cm(14))
context = { "date": "2022/11/30",
            "teachers": ["陳老師", "江老師", "張老師"],
            "students_data": students,
            "scores_chart": img }

doc.render(context)
doc.save(output_file)

wdFormatPDF = 17
app = Dispatch("Word.Application")
app.Visible = 1
app.DisplayAlerts = 0
docx = app.Documents.Open(os.getcwd()+"/"+output_file)
docx.SaveAs(os.getcwd()+"/"+output_pdf, 
            FileFormat=wdFormatPDF)
docx.Close()
app.Quit()


