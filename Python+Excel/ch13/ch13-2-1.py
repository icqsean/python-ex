from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb.active
ws["E1"] = "成績總分"
ws["E2"] = "=SUM(B2:D2)"
ws["E3"] = "=SUM(B3:D3)"
ws["E4"] = "=SUM(B4:D4)"
wb.save("成績管理_SUM.xlsx")   
wb.close()


