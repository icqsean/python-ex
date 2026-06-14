from pathlib import Path
from openpyxl import Workbook, load_workbook

wb_output = Workbook() 
ws_output = wb_output.active 
curr_row = 1
# 新增標題列
ws_output.cell(curr_row,1).value = "部門"
ws_output.cell(curr_row,2).value = "分類"
ws_output.cell(curr_row,3).value = "項目"
ws_output.cell(curr_row,4).value = "數量"
ws_output.cell(curr_row,5).value = "金額"
curr_row = curr_row + 1
# 彙整文具商品採購清單和計算小計
path = Path("文具商品採購/")
for item in path.iterdir():
    if item.match("*.xlsx"):
        wb = load_workbook(item)
        ws = wb.active
        department = ws["B2"].value
        for row in range(4, ws.max_row + 1):
            ws_output.cell(curr_row,1).value = department             # 部門
            ws_output.cell(curr_row,2).value = ws["A"+str(row)].value # 分類                    
            ws_output.cell(curr_row,3).value = ws["B"+str(row)].value # 項目
            price = ws["D"+str(row)].value
            quantity = ws["E"+str(row)].value
            ws_output.cell(curr_row,4).value = quantity               # 數量
            ws_output.cell(curr_row,5).value = price*quantity         # 金額
            curr_row = curr_row + 1
        wb.close()    

wb_output.save("文具商品採購清單.xlsx")
wb_output.close()
