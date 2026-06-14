from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb["Sheet"]
cols =["B", "C", "D"]
for col in cols:
    ws[col+"5"] = "=SUM("+col+"2:"+col+"4)"

wb.save("成績管理_SUM3.xlsx")   
wb.close()
