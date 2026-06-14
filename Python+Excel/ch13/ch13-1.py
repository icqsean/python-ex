from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb.active

cols = ["B", "C", "D"]
ws["E1"] = "成績總分"
for idx in range(2, 5):
    total = 0
    for col in cols:
        total = total + ws[col+str(idx)].value
    ws["E"+str(idx)] = total

wb.save("成績管理5.xlsx")
wb.close()