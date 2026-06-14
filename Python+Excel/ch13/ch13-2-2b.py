from openpyxl import load_workbook

wb = load_workbook("全班成績管理.xlsx")
ws = wb["A班"]  
ws["A6"] = "計數="
ws["B6"] = "= COUNT(B2:D5)"

wb.save("全班成績管理_COUNT.xlsx")
wb.close()