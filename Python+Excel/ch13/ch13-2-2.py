from openpyxl import load_workbook

wb = load_workbook("全班成績管理.xlsx")
ws = wb["A班"] 
ws["F1"] = "調整分數"
for i in range(2, 6):
    ws["F"+str(i)] = "=PRODUCT(D"+str(i)+":E"+str(i)+")"

wb.save("全班成績管理_PRODUCT.xlsx")
wb.close()