from openpyxl import load_workbook

wb = load_workbook("全班成績管理.xlsx")
ws = wb["A班"]  
ws["A6"] = "條件計數="
ws["B6"] = '= COUNTIF(B2:D5, ">70")'

wb.save("全班成績管理_COUNTIF2.xlsx")
wb.close()
