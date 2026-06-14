from openpyxl import load_workbook

wb = load_workbook("全班成績管理.xlsx")
ws = wb["A班"] 
ws["F1"] = "成績平均"
for i in range(2, 6):
    ws["F"+str(i)] = "=AVERAGE(B"+str(i)+":D"+str(i)+")"

wb.save("全班成績管理_AVERAGE.xlsx")   
wb.close()
