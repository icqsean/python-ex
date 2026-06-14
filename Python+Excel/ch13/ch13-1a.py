from openpyxl import load_workbook

wb = load_workbook("成績管理5.xlsx")
ws = wb.active

cols = ["B", "C", "D"]
ws["A5"] = "成績總分"
ws["A6"] = "成績平均"
for col in cols:
    total = 0
    for idx in range(2, 5):
        total = total + ws[col+str(idx)].value
    ws[col+"5"] = total
    ws[col+"6"] = total / 3

wb.save("成績管理6.xlsx")
wb.close()