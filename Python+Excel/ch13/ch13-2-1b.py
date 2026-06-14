from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb["Sheet"]
ws["E1"] = "成績總分"
for i in range(2, 5):
    ws["E"+str(i)] = "=SUM(B"+str(i)+":D"+str(i)+")"

wb.save("成績管理_SUM2.xlsx")   
wb.close()
