from openpyxl import load_workbook

wb = load_workbook("成績管理3.xlsx")
ws = wb.active

ws["B4"] = 75
ws["D4"] = 66

ws.cell(row=2, column=4).value = 82

wb.save("成績管理4.xlsx")
wb.close() 