from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb["Sheet"]
for col in ws.iter_cols():
    for cell in col:
        print(cell.value, end=" ")
    print()
wb.close()
