from openpyxl import load_workbook

wb = load_workbook("成績管理.xlsx")
ws = wb.active

ws["D1"] = "數學"
ws["D2"] = 80
ws["D3"] = 76

wb.save("成績管理2.xlsx")
wb.close() 