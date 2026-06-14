from openpyxl import load_workbook

wb = load_workbook("各班的成績管理.xlsx")
ws = wb.create_sheet(title="C班")

records = [("姓名", "國文", "英文"),
           ("張三", 78, 66),
           ("李四", 88, 85) ]
for item in records:
    ws.append(item)

wb.save("各班的成績管理2.xlsx")
wb.close()
