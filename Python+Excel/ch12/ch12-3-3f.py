from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb["Sheet"]
for col in ws.iter_cols(min_row=1,
                        max_row=2,
                        min_col=1,
                        max_col=3):
    for cell in col:
        print(cell.value, end=" ")
    print()     
wb.close()
