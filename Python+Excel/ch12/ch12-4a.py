from openpyxl import load_workbook

wb = load_workbook("各班的成績管理2.xlsx")
ws = wb.create_sheet(title="B班", index=1)

records = [("姓名", "國文", "英文"),
           ("王美麗", 68, 55)]
for item in records:
    ws.append(item)

wb.save("各班的成績管理3.xlsx")
wb.close()

