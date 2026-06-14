from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

wb = load_workbook("成績管理.xlsx")
ws = wb.active

print("索引1的字母:", get_column_letter(1))
print("字母C的索引:", column_index_from_string("C"))
print("最小欄字母:", get_column_letter(ws.min_column))
print("最大欄字母:", get_column_letter(ws.max_column))
wb.close()
