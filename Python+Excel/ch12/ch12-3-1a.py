from openpyxl import load_workbook

wb = load_workbook("成績管理2.xlsx")
ws = wb.active

data = { "A4": "王陽明",
         "B4": 65,
         "C4": 66,
         "D4": 55 }

for key, value in data.items():
    ws[key] = value

wb.save("成績管理3.xlsx")
wb.close()