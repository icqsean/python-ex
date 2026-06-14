from openpyxl import load_workbook

wb = load_workbook("各班的成績管理4.xlsx")
print("工作表數=", len(wb.sheetnames))
print(wb.sheetnames[0])
print(wb.sheetnames[1])
for ws_name in wb.sheetnames:
    print(ws_name, end="")
wb.close()
