from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb["Sheet"]
cells = ws["A1:D3"]
for i1, i2, i3, i4 in cells:
    print(i1.value, i2.value,
          i3.value, i4.value)
wb.close()
