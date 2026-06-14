from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb["Sheet"]
for row in ws.iter_rows():
    for cell in row:
        print(cell.value, end=" ")
    print()  
wb.close()
