from openpyxl import load_workbook

wb = load_workbook("成績管理.xlsx")
ws = wb.active

print("工作表名稱:", ws.title)
print("最小欄索引:", ws.min_column)
print("最大欄索引:", ws.max_column)
print("最小列索引:", ws.min_row)
print("最大列索引:", ws.max_row)
print("工作表尺寸:", ws.dimensions)
wb.close()
