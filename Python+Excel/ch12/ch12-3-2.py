from openpyxl import load_workbook

wb = load_workbook("成績管理4.xlsx")
ws = wb["Sheet"]
v  = ws["A1"].value
print("A1儲存格:", v)
v  = ws["B2"].value
print("B2儲存格:", v)
v  = ws.cell(row=3, column=3).value
print("C3儲存格:", v)
v  = ws.cell(row=3, column=4).value
print("D3儲存格:", v)
wb.close()
